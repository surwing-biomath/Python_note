# openpyxl模块是用于处理Microsoft Excel文件的第三方库
# 可以对Excel文件中的数据进行写入和读取

# 函数/属性名称                  # 功能描述
# load_workbook(filename)       # 打开已存在的表格，结果为工作簿对象
# workbook.sheetnames           # 工作簿对象的sheetnames属性，用于获取所有工作表的名称，结果为列表类型
# sheet.append(lst)             # 向工作表中添加一行数据，新数据接在工作表已有数据的后面
# workbook.save(excelname)      # 保存工作簿对象，参数为文件名
# Workbook()                    # 创建一个新的工作簿对象 

import weather
import openpyxl

html = weather.get_html()
lst = weather.parse_html(html)
workbook = openpyxl.Workbook()                    # 创建一个新的工作簿对象
sheet = workbook.create_sheet('天气预报')          # 创建一个新的工作表对象，参数为工作表名称
for item in lst:
    sheet.append(item)                            # 向工作表中添加一行数据，新数据接在工作表已有数据的后面

workbook.save('weather.xlsx')                     # 保存工作簿对象，参数为文件名